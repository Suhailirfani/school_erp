"""
HOD (Head of Department) Views
Full access to all management features
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
import json
import openpyxl
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa

from hadiya.models import (
    CustomUser, Staff, Course, Subject, Student, Attendance, Attendance_Report,
    Student_Result, Student_leave, Student_Feedback, Staff_Feedback,
    Student_Notification, Staff_Notification, SubjectType, Examination,
    Staff_leave, Enquiry, News
)


def hod_required(view_func):
    """Decorator to ensure only HOD can access views"""
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.user_type != 1:
            messages.error(request, 'Access denied. HOD privileges required.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper


@hod_required
def dashboard(request):
    """
    HOD Dashboard with statistics and overview
    """
    # Get counts
    students_count = Student.objects.count()
    staff_count = Staff.objects.count()
    courses_count = Course.objects.count()
    subjects_count = Subject.objects.count()
    
    # Get recent data
    recent_students = Student.objects.select_related('admin', 'course_id').order_by('-created_at')[:5]
    recent_enquiries = Enquiry.objects.select_related('course').order_by('-created_at')[:5]
    recent_news = News.objects.order_by('-created_at')[:3]
    
    # Get pending leave requests
    pending_student_leaves = Student_leave.objects.filter(status=0).count()
    pending_staff_leaves = Staff_leave.objects.filter(status=0).count()
    
    # Get unread feedbacks
    unread_student_feedbacks = Student_Feedback.objects.filter(feedback_reply__isnull=True).count()
    unread_staff_feedbacks = Staff_Feedback.objects.filter(feedback_reply__isnull=True).count()
    
    context = {
        'students_count': students_count,
        'staff_count': staff_count,
        'courses_count': courses_count,
        'subjects_count': subjects_count,
        'recent_students': recent_students,
        'recent_enquiries': recent_enquiries,
        'recent_news': recent_news,
        'pending_student_leaves': pending_student_leaves,
        'pending_staff_leaves': pending_staff_leaves,
        'unread_student_feedbacks': unread_student_feedbacks,
        'unread_staff_feedbacks': unread_staff_feedbacks,
    }
    
    return render(request, 'Hod/home_new.html', context)


@hod_required
def add_student(request):
    """Add new student"""
    courses = Course.objects.all()
    courses = Course.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'bulk_upload':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, 'Please select an excel file to upload')
                return redirect('add_student')
            
            try:
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Iterate rows skipping header
                for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                    try:
                        # Expected columns: First Name, Last Name, Username, Email, Password, Gender, Address, Course Name, Session Start, Session End, Admission Type
                        first_name = row[0]
                        last_name = row[1]
                        username = row[2]
                        email = row[3]
                        password = str(row[4]) if row[4] else '123456'
                        gender = row[5]
                        address = row[6]
                        course_name = row[7]
                        session_start = str(row[8])
                        session_end = str(row[9])
                        admission_type = row[10]
                        
                        # Validate mandatory fields
                        if not all([first_name, username, email, course_name]):
                            error_count += 1
                            errors.append(f"Row {i+2}: Missing required fields")
                            continue
                            
                        # Check existance
                        if CustomUser.objects.filter(email=email).exists():
                            error_count += 1
                            errors.append(f"Row {i+2}: Email {email} already exists")
                            continue
                        if CustomUser.objects.filter(username=username).exists():
                             error_count += 1
                             errors.append(f"Row {i+2}: Username {username} already exists")
                             continue
                             
                        # Get FKs
                        try:
                            course = Course.objects.get(name__iexact=course_name)
                        except Course.DoesNotExist:
                            error_count += 1
                            errors.append(f"Row {i+2}: Course '{course_name}' not found")
                            continue
                            

                        
                        # Create User
                        user = CustomUser.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            user_type=3
                        )
                        user.save()
                        
                        # Create Student
                        Student.objects.create(
                            admin=user,
                            address=address if address else '',
                            gender=gender if gender else 'Male',
                            course_id=course,
                            # session_year_id=session,
                            admission_type=admission_type if admission_type else 'Day Scholar',
                            uses_hostel=(admission_type == 'Hostel')
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {i+2}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'{success_count} Students uploaded successfully!')
                
                if error_count > 0:
                    messages.warning(request, f'{error_count} rows failed. Check errors below.')
                    for err in errors[:5]: # Show first 5 errors
                        messages.error(request, err)
                        
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                
            return redirect('add_student')
            
        else:
            # Manual Addition
            # Get form data
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            address = request.POST.get('address')
            gender = request.POST.get('gender')
            course_id = request.POST.get('course_id')
            course_id = request.POST.get('course_id')
            admission_type = request.POST.get('admission_type')
            profile_pic = request.FILES.get('profile_pic')
            phone_number = request.POST.get('phone_number')
            
            try:
                # Create user account
                user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    user_type=3  # Student
                )
                
                if profile_pic:
                    user.profile_pic = profile_pic
                user.save()
                
                # Create student profile
                course = Course.objects.get(id=course_id)
                
                student = Student.objects.create(
                    admin=user,
                    address=address,
                    gender=gender,
                    course_id=course,
                    # session_year_id=session,
                    admission_type=admission_type,
                    uses_hostel=(admission_type == 'Hostel'),
                    phone_number=phone_number
                )
                
                messages.success(request, f'Student {first_name} {last_name} added successfully!')
                return redirect('view_students')
                
            except Exception as e:
                messages.error(request, f'Error adding student: {str(e)}')
    
    context = {
        'courses': courses,
        'courses': courses,
    }
    return render(request, 'Hod/add_student.html', context)


@hod_required
def view_students(request):
    """View all students"""
def view_students(request):
    """View all students"""
    students = Student.objects.select_related('admin', 'course_id').all()
    
    # Filter by course if specified
    course_filter = request.GET.get('course')
    if course_filter:
        students = students.filter(course_id__id=course_filter)
    
    # Filter by session if specified
    session_filter = request.GET.get('session')
    if session_filter:
        students = students.filter(session_year_id__id=session_filter)
    
    # Search
    search_query = request.GET.get('search')
    if search_query:
        students = students.filter(
            Q(admin__first_name__icontains=search_query) |
            Q(admin__last_name__icontains=search_query) |
            Q(admin__email__icontains=search_query)
        )
    
    courses = Course.objects.all()

    
    context = {
        'students': students,
        'courses': courses,
    }
    return render(request, 'Hod/view_student.html', context)


@hod_required
def edit_student(request, student_id):
    """Edit student details"""
    student = get_object_or_404(Student, id=student_id)
    courses = Course.objects.all()

    
    if request.method == 'POST':
        try:
            # Update user info
            student.admin.first_name = request.POST.get('first_name')
            student.admin.last_name = request.POST.get('last_name')
            student.admin.email = request.POST.get('email')
            
            if request.FILES.get('profile_pic'):
                student.admin.profile_pic = request.FILES.get('profile_pic')
            
            password = request.POST.get('password')
            if password:
                student.admin.set_password(password)
            
            student.admin.save()
            
            # Update student info
            student.address = request.POST.get('address')
            student.gender = request.POST.get('gender')
            student.phone_number = request.POST.get('phone_number')
            student.course_id = Course.objects.get(id=request.POST.get('course_id'))
            # student.session_year_id = Session_year.objects.get(id=request.POST.get('session_id'))
            student.admission_type = request.POST.get('admission_type')
            student.uses_hostel = (student.admission_type == 'Hostel')
            student.save()
            
            messages.success(request, 'Student updated successfully!')
            return redirect('view_students')
            
        except Exception as e:
            messages.error(request, f'Error updating student: {str(e)}')
    
    context = {
        'student': student,
        'courses': courses,
    }
    return render(request, 'Hod/edit_student.html', context)


@hod_required
def delete_student(request, student_id):
    """Delete student"""
    student = get_object_or_404(Student, id=student_id)
    
    try:
        user = student.admin
        student.delete()
        user.delete()
        messages.success(request, 'Student deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting student: {str(e)}')
    
    return redirect('view_students')


@hod_required
def add_staff(request):
    """Add new staff member"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'bulk_upload':
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, 'Please select an excel file to upload')
                return redirect('add_staff')
                
            try:
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Iterate rows skipping header
                for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                    try:
                        # Expected columns: First Name, Last Name, Username, Email, Password, Gender, Address
                        first_name = row[0]
                        last_name = row[1]
                        username = row[2]
                        email = row[3]
                        password = str(row[4]) if row[4] else '123456'
                        gender = row[5]
                        address = row[6]
                        
                        # Validate mandatory fields
                        if not all([first_name, username, email]):
                            error_count += 1
                            errors.append(f"Row {i+2}: Missing required fields")
                            continue
                            
                        # Check existance
                        if CustomUser.objects.filter(email=email).exists():
                            error_count += 1
                            errors.append(f"Row {i+2}: Email {email} already exists")
                            continue
                        if CustomUser.objects.filter(username=username).exists():
                             error_count += 1
                             errors.append(f"Row {i+2}: Username {username} already exists")
                             continue
                        
                        # Create User
                        user = CustomUser.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                            user_type=2 # Staff
                        )
                        user.save()
                        
                        # Staff profile created by signal, just update it
                        staff = user.staff
                        staff.address = address if address else ''
                        staff.gender = gender if gender else 'Male'
                        staff.save()
                        
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {i+2}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'{success_count} Staff members uploaded successfully!')
                
                if error_count > 0:
                    messages.warning(request, f'{error_count} rows failed. Check errors below.')
                    for err in errors[:5]:
                        messages.error(request, err)
                        
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
            
            return redirect('add_staff')
            
        else:
            try:
                # Get form data
                first_name = request.POST.get('first_name')
                last_name = request.POST.get('last_name')
                username = request.POST.get('username')
                email = request.POST.get('email')
                password = request.POST.get('password')
                address = request.POST.get('address')
                gender = request.POST.get('gender')
                profile_pic = request.FILES.get('profile_pic')
                
                # Check if email or username already exists
                if CustomUser.objects.filter(email=email).exists():
                    messages.error(request, 'Email already exists!')
                    return redirect('add_staff')
                if CustomUser.objects.filter(username=username).exists():
                    messages.error(request, 'Username already exists!')
                    return redirect('add_staff')
                
                # Create user account
                user = CustomUser.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    user_type=2  # Staff
                )
                
                if profile_pic:
                    user.profile_pic = profile_pic
                user.save()
                
                # Update staff profile (automatically created by signal)
                staff = user.staff
                staff.address = address
                staff.gender = gender
                staff.save()
                
                messages.success(request, f'Staff {first_name} {last_name} added successfully!')
                return redirect('view_staff')
                
            except Exception as e:
                messages.error(request, f'Error adding staff: {str(e)}')
    
    return render(request, 'Hod/add_staff.html')


@hod_required
def view_staff(request):
    """View all staff members"""
    staff_members = Staff.objects.select_related('admin').all()
    
    # Search
    search_query = request.GET.get('search')
    if search_query:
        staff_members = staff_members.filter(
            Q(admin__first_name__icontains=search_query) |
            Q(admin__last_name__icontains=search_query) |
            Q(admin__email__icontains=search_query)
        )
    
    context = {
        'staff_members': staff_members
    }
    return render(request, 'Hod/view_staff.html', context)


@hod_required
def edit_staff(request, staff_id):
    """Edit staff details"""
    staff = get_object_or_404(Staff, id=staff_id)
    
    if request.method == 'POST':
        try:
            # Update user info
            staff.admin.first_name = request.POST.get('first_name')
            staff.admin.last_name = request.POST.get('last_name')
            staff.admin.email = request.POST.get('email')
            staff.admin.username = request.POST.get('username')
            
            if request.FILES.get('profile_pic'):
                staff.admin.profile_pic = request.FILES.get('profile_pic')
            
            password = request.POST.get('password')
            if password:
                staff.admin.set_password(password)
            
            staff.admin.save()
            
            # Update staff info
            staff.address = request.POST.get('address')
            staff.gender = request.POST.get('gender')
            staff.save()
            
            messages.success(request, 'Staff updated successfully!')
            return redirect('view_staff')
            
        except Exception as e:
            messages.error(request, f'Error updating staff: {str(e)}')
    
    context = {
        'staff': staff
    }
    return render(request, 'Hod/edit_staff.html', context)


@hod_required
def delete_staff(request, staff_id):
    """Delete staff member"""
    staff = get_object_or_404(Staff, id=staff_id)
    
    try:
        user = staff.admin
        staff.delete()
        user.delete()
        messages.success(request, 'Staff deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting staff: {str(e)}')
    
    return redirect('view_staff')


@hod_required
def add_course(request):
    """Add new course"""
    if request.method == 'POST':
        course_name = request.POST.get('course_name')
        try:
            course = Course(name=course_name)
            course.save()
            messages.success(request, 'Course added successfully!')
            return redirect('view_courses')
        except Exception as e:
            messages.error(request, f'Error adding course: {str(e)}')
    return render(request, 'Hod/add_course.html')


@hod_required
def view_courses(request):
    """View all courses"""
    courses = Course.objects.all()
    context = {'courses': courses}
    return render(request, 'Hod/view_course.html', context)


@hod_required
def edit_course(request, course_id):
    """Edit course details"""
    course = get_object_or_404(Course, id=course_id)
    if request.method == 'POST':
        try:
            course.name = request.POST.get('course_name')
            course.save()
            messages.success(request, 'Course updated successfully!')
            return redirect('view_courses')
        except Exception as e:
            messages.error(request, f'Error updating course: {str(e)}')
    context = {'course': course}
    return render(request, 'Hod/edit_course.html', context)


@hod_required
def delete_course(request, course_id):
    """Delete course"""
    course = get_object_or_404(Course, id=course_id)
    try:
        course.delete()
        messages.success(request, 'Course deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting course: {str(e)}')
    return redirect('view_courses')


@hod_required
def add_subject(request):
    """Add new subject"""
    courses = Course.objects.all()
    staffs = Staff.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'bulk_upload':
            excel_file = request.FILES.get('subject_file')
            if not excel_file:
                messages.error(request, 'Please select an excel file to upload')
                return redirect('add_subject')
                
            try:
                wb = openpyxl.load_workbook(excel_file)
                worksheet = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Iterate rows skipping header
                for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                    try:
                        # Expected columns: Subject Name, Course Name, Staff Email (Optional)
                        subject_name = row[0]
                        course_name = row[1]
                        staff_email = row[2] if len(row) > 2 else None
                        
                        if not subject_name or not course_name:
                            error_count += 1
                            errors.append(f"Row {i+2}: Missing Name or Course")
                            continue
                            
                        # Get Course
                        try:
                            course = Course.objects.get(name__iexact=course_name)
                        except Course.DoesNotExist:
                            error_count += 1
                            errors.append(f"Row {i+2}: Course '{course_name}' not found")
                            continue
                            
                        # Get Staff (Optional)
                        staff = None
                        if staff_email:
                            try:
                                staff = Staff.objects.get(admin__email__iexact=staff_email)
                            except Staff.DoesNotExist:
                                # We can choose to fail or just ignore the staff assignment
                                # Let's warn but proceed with None, or fail? 
                                # User requirement doesn't specify data strictness. 
                                # Best to error if provided but invalid to avoid data issues.
                                error_count += 1
                                errors.append(f"Row {i+2}: Staff email '{staff_email}' not found")
                                continue
                        
                        # Create Subject
                        Subject.objects.create(
                            name=subject_name,
                            course=course,
                            staff=staff
                        )
                        success_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {i+2}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'{success_count} Subjects uploaded successfully!')
                
                if error_count > 0:
                    messages.warning(request, f'{error_count} rows failed. Check errors below.')
                    for err in errors[:5]:
                        messages.error(request, err)
                        
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                
            return redirect('add_subject')
            
        else:
            # Single Add
            subject_name = request.POST.get('subject_name')
            course_id = request.POST.get('course_id')
            staff_id = request.POST.get('staff_id')
            subject_type_id = request.POST.get('subject_type_id')
            
            try:
                course = Course.objects.get(id=course_id)
                staff = Staff.objects.get(id=staff_id) if staff_id else None
                subject_type = SubjectType.objects.get(id=subject_type_id) if subject_type_id else None
                
                subject = Subject(
                    name=subject_name,
                    course=course,
                    staff=staff,
                    subject_type=subject_type,
                    max_te_marks=request.POST.get('max_te_marks', 100),
                    max_ce_marks=request.POST.get('max_ce_marks', 100)
                )
                subject.save()
                messages.success(request, 'Subject added successfully!')
                return redirect('view_subjects')
            except Exception as e:
                messages.error(request, f'Error adding subject: {str(e)}')
            
    subject_types = SubjectType.objects.all()
    context = {
        'courses': courses,
        'staffs': staffs,
        'subject_types': subject_types
    }
    return render(request, 'Hod/add_subject.html', context)


@hod_required
def view_subjects(request):
    """View all subjects"""
    subjects = Subject.objects.select_related('course', 'staff').all()
    context = {'subjects': subjects}
    return render(request, 'Hod/view_subject.html', context)


@hod_required
def edit_subject(request, subject_id):
    """Edit subject details"""
    subject = get_object_or_404(Subject, id=subject_id)
    courses = Course.objects.all()
    staffs = Staff.objects.all()
    
    if request.method == 'POST':
        try:
            subject.name = request.POST.get('subject_name')
            course_id = request.POST.get('course_id')
            staff_id = request.POST.get('staff_id')
            subject_type_id = request.POST.get('subject_type_id')
            
            subject.course = Course.objects.get(id=course_id)
            if staff_id:
                subject.staff = Staff.objects.get(id=staff_id)
            else:
                subject.staff = None
                
            if subject_type_id:
                subject.subject_type = SubjectType.objects.get(id=subject_type_id)
            else:
                subject.subject_type = None
                
            subject.max_te_marks = request.POST.get('max_te_marks', 100)
            subject.max_ce_marks = request.POST.get('max_ce_marks', 100)
                
            subject.save()
            messages.success(request, 'Subject updated successfully!')
            return redirect('view_subjects')
        except Exception as e:
            messages.error(request, f'Error updating subject: {str(e)}')
            
    subject_types = SubjectType.objects.all()
    context = {
        'subject': subject,
        'courses': courses,
        'staffs': staffs,
        'subject_types': subject_types
    }
    return render(request, 'Hod/edit_subject.html', context)


@hod_required
def delete_subject(request, subject_id):
    """Delete subject"""
    subject = get_object_or_404(Subject, id=subject_id)
    try:
        subject.delete()
        messages.success(request, 'Subject deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting subject: {str(e)}')
    return redirect('view_subjects')





@hod_required
def admin_view_attendance(request):
    """View to select criteria for viewing attendance"""
    courses = Course.objects.all()
    students = Student.objects.select_related('admin').all()

    
    context = {
        'courses': courses,
        'students': students,
    }
    return render(request, 'Hod/view_attendance.html', context)


@csrf_exempt
def admin_get_attendance_dates(request):
    """AJAX view to get dates for HOD"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
    
    course_id = request.POST.get("course_id")
    try:
        course = Course.objects.get(id=course_id)
        
        attendance_objects = Attendance.objects.filter(course=course)
        
        list_data = []
        for attendance in attendance_objects:
            data_small = {
                "id": attendance.id,
                "attendance_date": str(attendance.attendance_date),
            }
            list_data.append(data_small)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@csrf_exempt
def admin_get_attendance_student(request):
    """AJAX view to get student attendance data for HOD view"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    attendance_date = request.POST.get("attendance_date")
    attendance_id = request.POST.get("attendance_id")
    
    try:
        attendance = Attendance.objects.get(id=attendance_id)
        attendance_data = Attendance_Report.objects.filter(attendance=attendance)
        
        list_data = []
        for student_data in attendance_data:
            data_small = {
                "id": student_data.student.admin.id,
                "name": student_data.student.admin.first_name + " " + student_data.student.admin.last_name,
                "status": 1 if student_data.is_present else 0,
                "leave_status": student_data.leave_status if student_data.leave_status else ''
            }
            list_data.append(data_small)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@hod_required
def download_sample_file(request, file_type):
    """Download sample excel file for bulk upload"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Data"
    
    if file_type == 'student':
        # Headers: First Name, Last Name, Username, Email, Password, Gender, Address, Course, Admission Type
        headers = ['First Name', 'Last Name', 'Username', 'Email', 'Password', 'Gender', 'Address', 'Course', 'Admission Type']
        ws.append(headers)
        # Add a sample row
        ws.append(['John', 'Doe', 'johndoe', 'john@example.com', '123456', 'Male', '123 Street', 'Computer Science', 'Day Scholar'])
        filename = "student_upload_sample.xlsx"
        
    elif file_type == 'staff':
        # Headers: First Name, Last Name, Username, Email, Password, Gender, Address
        headers = ['First Name', 'Last Name', 'Username', 'Email', 'Password', 'Gender', 'Address']
        ws.append(headers)
        # Add sample row
        ws.append(['Jane', 'Doe', 'janedoe', 'jane@example.com', '123456', 'Female', '456 Avenue'])
        filename = "staff_upload_sample.xlsx"
    
    else:
        return HttpResponse("Invalid file type")
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response


@csrf_exempt
def admin_get_attendance_report_data(request):
    """Get aggregated attendance report based on filters"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    course_id = request.POST.get("course_id")
    filter_type = request.POST.get("filter_type")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")
    month = request.POST.get("month")
    year = request.POST.get("year")
    
    try:
        # Determine Courses to Process
        if course_id:
            courses = Course.objects.filter(id=course_id)
        else:
            courses = Course.objects.all()
            
        final_data = [] # List of { course_name: "", data: [] }
        
        for course in courses:
            # Base query for Attendance
            attendance_query = Attendance.objects.filter(course=course)
            
            if filter_type == "day":
                attendance_query = attendance_query.filter(attendance_date=start_date)
            elif filter_type == "weekly":
                date_obj = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
                week_start = date_obj - datetime.timedelta(days=date_obj.weekday()) 
                week_end = week_start + datetime.timedelta(days=6)
                attendance_query = attendance_query.filter(attendance_date__range=[week_start, week_end])
            elif filter_type == "month":
                attendance_query = attendance_query.filter(attendance_date__month=month, attendance_date__year=year)
            elif filter_type == "year":
                attendance_query = attendance_query.filter(attendance_date__year=year)
            elif filter_type == "range":
                attendance_query = attendance_query.filter(attendance_date__range=[start_date, end_date])
            elif filter_type == "till_today":
                today = datetime.date.today()
                attendance_query = attendance_query.filter(attendance_date__lte=today)
                
            # LOGIC TO HANDLE DUPLICATES
            all_attendance = attendance_query.values('id', 'attendance_date', 'updated_at')
            
            latest_attendance_map = {}
            for att in all_attendance:
                d = str(att['attendance_date']) 
                if d not in latest_attendance_map:
                    latest_attendance_map[d] = att
                else:
                    if att['updated_at'] > latest_attendance_map[d]['updated_at']:
                        latest_attendance_map[d] = att
                        
            valid_attendance_ids = [item['id'] for item in latest_attendance_map.values()]
            total_days = len(valid_attendance_ids)
            
            students = Student.objects.filter(course_id=course)
            
            course_student_data = []
            if students.exists():
                for student in students:
                    present_count = Attendance_Report.objects.filter(
                        student=student, 
                        attendance__id__in=valid_attendance_ids, 
                        is_present=True
                    ).count()
                    
                    absent_count = Attendance_Report.objects.filter(
                        student=student, 
                        attendance__id__in=valid_attendance_ids, 
                        is_present=False
                    ).count()
                    
                    percentage = (present_count / total_days * 100) if total_days > 0 else 0
                    
                    data = {
                        "name": f"{student.admin.first_name} {student.admin.last_name}",
                        "total_days": total_days,
                        "present": present_count,
                        "absent": absent_count,
                        "percentage": round(percentage, 2)
                    }
                    course_student_data.append(data)
            
            # Append even if empty to show the header? Or strict checking. Let's append if course exists.
            final_data.append({
                "course_name": course.name,
                "data": course_student_data
            })
            
        return JsonResponse(json.dumps(final_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@csrf_exempt
def admin_get_student_attendance_data(request):
    """Get detailed attendance history for a single student"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
         
    student_id = request.POST.get("student_id")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")
    
    try:
        student = Student.objects.get(admin__id=student_id)
        
        # Get all reports
        reports = Attendance_Report.objects.filter(student=student)
        
        if start_date and end_date:
            reports = reports.filter(attendance__attendance_date__range=[start_date, end_date])
            
        # Select related for performance access
        reports = reports.select_related('attendance').order_by('-attendance__attendance_date')
        
        # LOGIC TO HANDLE DUPLICATES: Show only the latest record for each date
        processed_dates = set()
        list_data = []
        
        # Since we ordered by date desc, we might still have duplicates with same date.
        # We need to sort by date desc AND updated_at desc to ensure first encountered is latest.
        # But Django can't sort by related updated_at easily without annotation if not in model ordering.
        # Let's fetch and process in python.
        
        all_reports = []
        for r in reports:
            all_reports.append({
                'date': str(r.attendance.attendance_date),
                'updated_at': r.attendance.updated_at, 
                'obj': r
            })
            
        # Sort by date desc, then updated_at desc
        all_reports.sort(key=lambda x: (x['date'], x['updated_at']), reverse=True)
        
        for item in all_reports:
            date_str = item['date']
            if date_str not in processed_dates:
                processed_dates.add(date_str)
                report = item['obj']
                
                data = {
                    "date": date_str,
                    "status": "Present" if report.is_present else "Absent",
                    "leave_status": report.leave_status if report.leave_status else "-"
                }
                list_data.append(data)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@csrf_exempt
def admin_analyze_attendance(request):
    """Analyze attendance for Exam Eligibility or Awards"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
         
    course_id = request.POST.get("course_id")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")
    analysis_type = request.POST.get("analysis_type") # 'below' or 'above'
    threshold = float(request.POST.get("threshold", 0))
    
    try:
        # Determine Courses to Process
        if course_id:
            courses = Course.objects.filter(id=course_id)
        else:
            courses = Course.objects.all()
            
        final_data = []
        
        for course in courses:
            students = Student.objects.filter(course_id=course)
            
            # Base Attendance Query
            attendance_query = Attendance.objects.filter(course=course)
            if start_date and end_date:
                attendance_query = attendance_query.filter(attendance_date__range=[start_date, end_date])
                
            # Deduplication Logic (Same as Report Data)
            all_attendance = attendance_query.values('id', 'attendance_date', 'updated_at')
            latest_attendance_map = {}
            for att in all_attendance:
                d = str(att['attendance_date'])
                if d not in latest_attendance_map:
                    latest_attendance_map[d] = att
                else:
                    if att['updated_at'] > latest_attendance_map[d]['updated_at']:
                        latest_attendance_map[d] = att
            
            valid_attendance_ids = [item['id'] for item in latest_attendance_map.values()]
            total_days = len(valid_attendance_ids)
            
            course_list_data = []
            if total_days > 0 and students.exists():
                for student in students:
                    present_count = Attendance_Report.objects.filter(
                        student=student, 
                        attendance__id__in=valid_attendance_ids, 
                        is_present=True
                    ).count()
                    
                    percentage = (present_count / total_days * 100)
                    percentage = round(percentage, 2)
                    
                    # Filtering based on Criteria
                    include_record = False
                    status = ""
                    
                    if analysis_type == 'below':
                        if percentage < threshold:
                            include_record = True
                            status = "Correction Required"
                    elif analysis_type == 'above':
                        if percentage >= threshold:
                            include_record = True
                            status = "Eligible for Award"
                    
                    if include_record:
                        data = {
                            "name": f"{student.admin.first_name} {student.admin.last_name}",
                            "total_days": total_days,
                            "present": present_count,
                            "percentage": percentage,
                            "status": status
                        }
                        course_list_data.append(data)
                        
                # Sorting
                if analysis_type == 'below':
                    course_list_data.sort(key=lambda x: x['percentage'])
                else:
                    course_list_data.sort(key=lambda x: x['percentage'], reverse=True)
            
            # Append if data exists or even if empty if desired (keeping empty for now to match structure)
            final_data.append({
                "course_name": course.name,
                "data": course_list_data
            })
            
        return JsonResponse(json.dumps(final_data), content_type="application/json", safe=False)
        
    except Exception as e:
        return HttpResponse(str(e))

@hod_required
def add_result(request):
    """View to add/edit student results"""
    courses = Course.objects.all()
    exams = Examination.objects.all()
    context = {
        'courses': courses,
        'exams': exams
    }
    return render(request, 'Hod/add_result.html', context)


@csrf_exempt
def admin_get_subjects(request):
    """Fetch subjects for a selected course"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
    
    course_id = request.POST.get("course_id")
    try:
        course = Course.objects.get(id=course_id)
        subjects = Subject.objects.filter(course=course)
        list_data = []
        for subject in subjects:
            data = {"id": subject.id, "name": subject.name}
            list_data.append(data)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@csrf_exempt
def admin_get_students_for_result(request):
    """Fetch students and their existing marks for a subject"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
         
    course_id = request.POST.get("course_id")
    subject_id = request.POST.get("subject_id")
    exam_id = request.POST.get("exam_id")
    
    try:
        course = Course.objects.get(id=course_id)
        subject = Subject.objects.get(id=subject_id)
        students = Student.objects.filter(course_id=course)
        
        if exam_id:
            exam = Examination.objects.get(id=exam_id)
        else:
            exam = None

        list_data = []
        for student in students:
            # Check if result already exists
            try:
                result = Student_Result.objects.get(student=student, subject=subject, exam=exam)
                ce_marks = result.ce_marks
                te_marks = result.te_marks
            except Student_Result.DoesNotExist:
                ce_marks = 0
                te_marks = 0
            
            data = {
                "id": student.admin.id, # Use User ID for identification
                "name": f"{student.admin.first_name} {student.admin.last_name}",
                "ce_marks": ce_marks,
                "te_marks": te_marks
            }
            list_data.append(data)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@hod_required
def manage_exam(request):
    """View to manage exams"""
    exams = Examination.objects.all()
    context = {
        'exams': exams
    }
    return render(request, 'Hod/manage_exam.html', context)

@hod_required
def add_exam(request):
    """View to add exam"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            date = request.POST.get('date')
            
            exam = Examination(name=name, date=date if date else None)
            exam.save()
            messages.success(request, "Examination Added Successfully!")
            return redirect('manage_exam')
        except Exception as e:
             messages.error(request, f"Failed to Add Examination: {e}")
             return redirect('manage_exam')
    return render(request, 'Hod/add_exam.html')

@hod_required
def delete_exam(request, exam_id):
    try:
        exam = Examination.objects.get(id=exam_id)
        exam.delete()
        messages.success(request, "Examination Deleted Successfully!")
    except Exception as e:
        messages.error(request, f"Failed to Delete Examination: {e}")
    return redirect('manage_exam')


@csrf_exempt
def admin_save_student_result(request):
    """Save student marks"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")

    try:
        student_ids = request.POST.getlist("student_ids[]")
        ce_marks_list = request.POST.getlist("ce_marks[]")
        te_marks_list = request.POST.getlist("te_marks[]")
        subject_id = request.POST.get("subject_id")
        exam_id = request.POST.get("exam_id")
        
        subject = Subject.objects.get(id=subject_id)
        # Use existing exam logic or None
        if exam_id:
             exam = Examination.objects.get(id=exam_id)
        else:
             exam = None
        
        for i in range(len(student_ids)):
            student_admin_id = student_ids[i]
            ce_marks = float(ce_marks_list[i])
            te_marks = float(te_marks_list[i])
            
            student = Student.objects.get(admin__id=student_admin_id)
            
            # Create or Update
            result, created = Student_Result.objects.update_or_create(
                student=student,
                subject=subject,
                exam=exam,
                defaults={
                    'ce_marks': ce_marks,
                    'te_marks': te_marks
                }
            )
            
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


@hod_required
def admin_view_result(request):
    """View Results Main Page"""
    courses = Course.objects.all()
    exams = Examination.objects.all()
    context = {
        'courses': courses,
        'exams': exams
    }
    return render(request, 'Hod/admin_view_result.html', context)


@csrf_exempt
def admin_get_students(request):
    """Fetch students for a course (AJAX)"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
         
    course_id = request.POST.get("course_id")
    try:
        course = Course.objects.get(id=course_id)
        students = Student.objects.filter(course_id=course)
        
        list_data = []
        for student in students:
            data = {
                "id": student.admin.id, 
                "name": f"{student.admin.first_name} {student.admin.last_name}"
            }
            list_data.append(data)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))


@csrf_exempt
def admin_get_result_data(request):
    """Fetch result data for display (AJAX)"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")

    view_type = request.POST.get("view_type")
    course_id = request.POST.get("course_id")
    exam_id = request.POST.get("exam_id")
    
    try:
        course = Course.objects.get(id=course_id)
        if exam_id:
             exam = Examination.objects.get(id=exam_id)
        else:
             exam = None
             
        list_data = []

        if view_type == 'class_wise':
            subject_id = request.POST.get("subject_id")
            subject = Subject.objects.get(id=subject_id)
            students = Student.objects.filter(course_id=course)

            for student in students:
                try:
                    result = Student_Result.objects.get(student=student, subject=subject, exam=exam)
                    ce = result.ce_marks
                    te = result.te_marks
                    total = ce + te
                except Student_Result.DoesNotExist:
                    ce = 0
                    te = 0
                    total = 0
                
                data = {
                    "name": f"{student.admin.first_name} {student.admin.last_name}",
                    "ce_marks": ce,
                    "te_marks": te,
                    "total_marks": total
                }
                list_data.append(data)

        elif view_type == 'student_wise':
            student_id = request.POST.get("student_id")
            student = Student.objects.get(admin__id=student_id)
            subjects = Subject.objects.filter(course=course)

            for subject in subjects:
                try:
                    result = Student_Result.objects.get(student=student, subject=subject, exam=exam)
                    ce = result.ce_marks
                    te = result.te_marks
                    total = ce + te
                except Student_Result.DoesNotExist:
                    ce = 0
                    te = 0
                    total = 0
                
                data = {
                    "subject_name": subject.name,
                    "max_ce_marks": subject.max_ce_marks,
                    "max_te_marks": subject.max_te_marks,
                    "ce_marks": ce,
                    "te_marks": te,
                    "total_marks": total
                }
                list_data.append(data)
        
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)

    except Exception as e:
        return HttpResponse(str(e))


def render_to_pdf(template_src, context_dict={}):
    """Helper to render HTML to PDF"""
    template = get_template(template_src)
    html  = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


@csrf_exempt
def generate_result_pdf(request):
    """Generate PDF for Results"""
    if request.method != "POST":
         return HttpResponse("Method Not Allowed")
    
    view_type = request.POST.get("pdf_view_type")
    course_id = request.POST.get("pdf_course_id")
    exam_id = request.POST.get("pdf_exam_id")
    
    try:
        course = Course.objects.get(id=course_id)
        if exam_id:
             exam = Examination.objects.get(id=exam_id)
        else:
             exam = None
             
        context = {
            'view_type': view_type,
            'course': course,
            'exam': exam
        }
        
        data = []
        if view_type == 'class_wise':
            subject_id = request.POST.get("pdf_subject_id")
            subject = Subject.objects.get(id=subject_id)
            students = Student.objects.filter(course_id=course)
            
            context['subject'] = subject
            
            for student in students:
                try:
                    result = Student_Result.objects.get(student=student, subject=subject, exam=exam)
                    ce = result.ce_marks
                    te = result.te_marks
                    total = ce + te
                except Student_Result.DoesNotExist:
                    ce = 0
                    te = 0
                    total = 0
                
                data.append({
                    "name": f"{student.admin.first_name} {student.admin.last_name}",
                    "ce_marks": ce,
                    "te_marks": te,
                    "total_marks": total
                })
            context['data'] = data

        elif view_type == 'student_wise':
            student_id = request.POST.get("pdf_student_id")
            student = Student.objects.get(admin__id=student_id)
            subjects = Subject.objects.filter(course=course)
            
            context['student'] = student
            
            total_obtained = 0
            total_max = 0
            
            for subject in subjects:
                try:
                    result = Student_Result.objects.get(student=student, subject=subject, exam=exam)
                    ce = result.ce_marks
                    te = result.te_marks
                    total = ce + te
                except Student_Result.DoesNotExist:
                    ce = 0
                    te = 0
                    total = 0
                
                sub_max = subject.max_ce_marks + subject.max_te_marks
                total_obtained += total
                total_max += sub_max
                
                # Percentage for grade
                perc = (total / sub_max * 100) if sub_max > 0 else 0
                
                data.append({
                    "subject_name": subject.name,
                    "max_ce_marks": subject.max_ce_marks,
                    "max_te_marks": subject.max_te_marks,
                    "ce_marks": ce,
                    "te_marks": te,
                    "total_marks": total,
                    "percentage": perc
                })
            context['data'] = data
            context['total_obtained'] = total_obtained
            context['total_max'] = total_max
            context['overall_percentage'] = (total_obtained / total_max * 100) if total_max > 0 else 0
            
        return render_to_pdf('Hod/result_pdf_template.html', context)
        
    except Exception as e:
        return HttpResponse(f"Error Generating PDF: {str(e)}")


@hod_required
def admin_view_student_leave(request):
    """View Pending Student Leave Requests"""
    leaves = Student_leave.objects.all().order_by('-id')
    context = {'leaves': leaves}
    return render(request, 'Hod/student_leave_view.html', context)


@hod_required
def admin_view_staff_leave(request):
    """View Pending Staff Leave Requests"""
    leaves = Staff_leave.objects.all().order_by('-id')
    context = {'leaves': leaves}
    return render(request, 'Hod/staff_leave_view.html', context)


@csrf_exempt
def admin_analyze_result(request):
    """Analyze student results logic"""
    if request.method == "POST":
        action = request.POST.get("action")
        course_id = request.POST.get("course_id")
        exam_id = request.POST.get("exam_id")
        subject_type_id = request.POST.get("subject_type_id")
        
        if action == 'fetch_graph':
            try:
                course = Course.objects.get(id=course_id)
                subjects = Subject.objects.filter(course=course)
                
                if subject_type_id:
                     subjects = subjects.filter(subject_type__id=subject_type_id)
                
                subject_labels = []
                pass_percentages = []
                avg_marks = []
                
                if exam_id:
                     exam = Examination.objects.get(id=exam_id)
                else:
                     exam = None

                for subject in subjects:
                    if exam:
                        results = Student_Result.objects.filter(subject=subject, exam=exam)
                    else:
                        results = Student_Result.objects.filter(subject=subject)
                        
                    total_students = results.count()
                    
                    if total_students >= 0: # Show even if 0 to maintain chart structure if needed, or filter? Let's keep it.
                        passed = 0
                        total_score = 0
                        max_score_sum = 0
                        
                        if total_students > 0:
                            for res in results:
                                 total = res.ce_marks + res.te_marks
                                 sub_max = subject.max_ce_marks + subject.max_te_marks
                                 if sub_max > 0 and (total/sub_max)*100 >= 40:
                                     passed += 1
                                 
                                 total_score += total
                                 max_score_sum += sub_max
                            
                            pass_perc = (passed / total_students) * 100
                            avg_perc = (total_score / max_score_sum * 100) if max_score_sum > 0 else 0
                        else:
                            pass_perc = 0
                            avg_perc = 0
                        
                        subject_labels.append(subject.name)
                        pass_percentages.append(round(pass_perc, 2))
                        avg_marks.append(round(avg_perc, 2))
                    
                data = {
                    "labels": subject_labels,
                    "pass_percentages": pass_percentages,
                    "avg_marks": avg_marks
                }
                return JsonResponse(json.dumps(data), content_type="application/json", safe=False)
                
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
                
    courses = Course.objects.all()
    exams = Examination.objects.all()
    subject_types = SubjectType.objects.all()
    context = {
        "courses": courses,
        "exams": exams,
        "subject_types": subject_types
    }
    return render(request, "Hod/analyze_result.html", context)


@csrf_exempt
def admin_approve_student_leave(request, leave_id):
    """Approve Student Leave"""
    leave = Student_leave.objects.get(id=leave_id)
    leave.status = 1
    leave.save()
    
    # Notify Student
    notification = Student_Notification(
        student=leave.student,
        message=f"Leave Approved for {leave.leave_date}"
    )
    notification.save()
    
    messages.success(request, "Student Leave Approved")
    return redirect('admin_view_student_leave')


@hod_required
def admin_disapprove_student_leave(request, leave_id):
    """Disapprove Student Leave"""
    leave = get_object_or_404(Student_leave, id=leave_id)
    leave.status = 2
    leave.admin_remark = request.POST.get('admin_remark', '') # Capture remark
    leave.save()
    messages.error(request, "Student Leave Rejected")
    return redirect('admin_view_student_leave')


@hod_required
def admin_take_attendance(request):
    """View to select Course/Session for taking attendance (HOD)"""
    courses = Course.objects.all()
    context = {
        'courses': courses,
    }
    return render(request, 'Hod/take_attendance.html', context)

@csrf_exempt
def admin_get_students_attendance(request):
    """AJAX view to get students based on course (HOD)"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    course_id = request.POST.get('course_id')
    
    try:
        course = Course.objects.get(id=course_id)
        students = Student.objects.filter(course_id=course)
        
        list_data = []
        for student in students:
            data_small = {
                "id": student.admin.id, 
                "name": student.admin.first_name + " " + student.admin.last_name
            }
            list_data.append(data_small)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))

@csrf_exempt
def admin_save_attendance_data(request):
    """AJAX view to save attendance data (HOD)"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    student_ids = request.POST.get("student_ids")
    course_id = request.POST.get("course_id")
    attendance_date = request.POST.get("attendance_date")
    
    try:
        course = Course.objects.get(id=course_id)
        
        # Save Attendance Record
        attendance = Attendance(
            course=course,
            attendance_date=attendance_date
        )
        attendance.save()
        
        # Save Student Attendance Reports
        json_student_ids = json.loads(student_ids)
        for stud_dict in json_student_ids:
            student_id = stud_dict['id']
            status = stud_dict['status'] 
            leave_status = stud_dict.get('leave_status', '')
            
            student = Student.objects.get(admin_id=student_id)
            
            attendance_report = Attendance_Report(
                student=student,
                attendance=attendance,
                is_present=bool(status),
                leave_status=leave_status
            )
            attendance_report.save()
            
        return HttpResponse("OK")
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")


@hod_required
def admin_update_attendance(request):
    """View to select date for updating attendance (HOD)"""
    courses = Course.objects.all()
    context = {
        'courses': courses,
    }
    return render(request, 'Hod/update_attendance.html', context)

@csrf_exempt
def admin_get_attendance_dates(request):
    """AJAX view to get dates with existing attendance records (HOD)"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
    
    course_id = request.POST.get("course_id")
    try:
        course = Course.objects.get(id=course_id)
        
        attendance_objects = Attendance.objects.filter(course=course)
        
        list_data = []
        for attendance in attendance_objects:
            data_small = {
                "id": attendance.id,
                "attendance_date": str(attendance.attendance_date),
            }
            list_data.append(data_small)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))

@csrf_exempt
def admin_get_attendance_student(request):
    """AJAX view to get student attendance data for update (HOD)"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    attendance_date = request.POST.get("attendance_date")
    attendance_id = request.POST.get("attendance_id")
    
    try:
        attendance = Attendance.objects.get(id=attendance_id)
        attendance_data = Attendance_Report.objects.filter(attendance=attendance)
        
        list_data = []
        for student_data in attendance_data:
            data_small = {
                "id": student_data.student.admin.id,
                "name": student_data.student.admin.first_name + " " + student_data.student.admin.last_name,
                "status": 1 if student_data.is_present else 0,
                "leave_status": student_data.leave_status if student_data.leave_status else ''
            }
            list_data.append(data_small)
            
        return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)
    except Exception as e:
        return HttpResponse(str(e))

@csrf_exempt
def admin_save_updateattendance_data(request):
    """AJAX view to save updated attendance data (HOD)"""
    if request.method != "POST":
        return HttpResponse("Method Not Allowed")
        
    student_ids = request.POST.get("student_ids")
    attendance_date = request.POST.get("attendance_date")
    attendance_id = request.POST.get("attendance_id")
    
    try:
        attendance = Attendance.objects.get(id=attendance_id)
        attendance.attendance_date = attendance_date
        attendance.save()
        
        json_student_ids = json.loads(student_ids)
        for stud_dict in json_student_ids:
            student_id = stud_dict['id']
            status = stud_dict['status']
            leave_status = stud_dict.get('leave_status', '')
            
            student = Student.objects.get(admin_id=student_id)
            
            attendance_report = Attendance_Report.objects.get(student=student, attendance=attendance)
            attendance_report.is_present = bool(status)
            attendance_report.leave_status = leave_status
            attendance_report.save()
            
        return HttpResponse("OK")
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")


    return redirect('admin_view_student_leave')


@hod_required
def admin_approve_staff_leave(request, leave_id):
    """Approve Staff Leave"""
    leave = get_object_or_404(Staff_leave, id=leave_id)
    leave.status = 1
    leave.admin_remark = request.POST.get('admin_remark', '')
    leave.save()
    messages.success(request, "Staff Leave Approved")
    return redirect('admin_view_staff_leave')


@hod_required
def admin_disapprove_staff_leave(request, leave_id):
    """Disapprove Staff Leave"""
    leave = get_object_or_404(Staff_leave, id=leave_id)
    leave.status = 2
    leave.admin_remark = request.POST.get('admin_remark', '')
    leave.save()
    messages.error(request, "Staff Leave Rejected")
    return redirect('admin_view_staff_leave')


@hod_required
def admin_view_student_feedback(request):
    """View Student Feedback"""
    feedbacks = Student_Feedback.objects.all().order_by('-id')
    context = {'feedbacks': feedbacks}
    return render(request, 'Hod/student_feedback_view.html', context)

@hod_required
def admin_reply_student_feedback(request):
    """Reply to Student Feedback"""
    if request.method != 'POST':
        return HttpResponse("Method Not Allowed")
        
    feedback_id = request.POST.get('feedback_id')
    reply_message = request.POST.get('reply_message')
    
    try:
        feedback = Student_Feedback.objects.get(id=feedback_id)
        feedback.feedback_reply = reply_message
        feedback.save()
        messages.success(request, "Reply Sent Successfully")
    except Exception as e:
        messages.error(request, f"Failed to Send Reply: {str(e)}")
        
    return redirect('admin_view_student_feedback')


@hod_required
def admin_view_staff_feedback(request):
    """View Staff Feedback"""
    feedbacks = Staff_Feedback.objects.all().order_by('-id')
    context = {'feedbacks': feedbacks}
    return render(request, 'Hod/staff_feedback_view.html', context)


@hod_required
def admin_reply_staff_feedback(request):
    """Reply to Staff Feedback"""
    if request.method != 'POST':
        return HttpResponse("Method Not Allowed")
        
    feedback_id = request.POST.get('feedback_id')
    reply_message = request.POST.get('reply_message')
    
    try:
        feedback = Staff_Feedback.objects.get(id=feedback_id)
        feedback.feedback_reply = reply_message
        feedback.save()
        messages.success(request, "Reply Sent Successfully")
    except Exception as e:
        messages.error(request, f"Failed to Send Reply: {str(e)}")
        
    return redirect('admin_view_staff_feedback')


@hod_required
def manage_subject_type(request):
    """Manage Subject Types"""
    subject_types = SubjectType.objects.all()
    context = {
        'subject_types': subject_types
    }
    return render(request, 'Hod/manage_subject_type.html', context)


@hod_required
def add_subject_type(request):
    """Add Subject Type"""
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            try:
                SubjectType.objects.create(name=name)
                messages.success(request, "Subject Type Added Successfully!")
                return redirect('manage_subject_type')
            except Exception as e:
                messages.error(request, f"Error: {e}")
        else:
            messages.error(request, "Name is required")
            
    return render(request, 'Hod/add_subject_type.html')


@hod_required
def edit_subject_type(request, subject_type_id):
    """Edit Subject Type"""
    subject_type = get_object_or_404(SubjectType, id=subject_type_id)
    
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            try:
                subject_type.name = name
                subject_type.save()
                messages.success(request, "Subject Type Updated Successfully!")
                return redirect('manage_subject_type')
            except Exception as e:
                messages.error(request, f"Error: {e}")
        else:
            messages.error(request, "Name is required")
            
    return render(request, 'Hod/edit_subject_type.html', {'subject_type': subject_type})


@hod_required
def delete_subject_type(request, subject_type_id):
    """Delete Subject Type"""
    try:
        subject_type = SubjectType.objects.get(id=subject_type_id)
        subject_type.delete()
        messages.success(request, "Subject Type Deleted Successfully!")
    except Exception as e:
        messages.error(request, f"Error: {e}")
    return redirect('manage_subject_type')